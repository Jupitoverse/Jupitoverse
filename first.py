import openpyxl
from openpyxl import Workbook
import os
import time

# Define file paths
master_file = 'C:/Users/abhisha3/Desktop/Projects/Excel/Master.xlsx'
child_files = [
    'C:/Users/abhisha3/Desktop/Projects/Excel/Abhi.xlsx',
    'C:/Users/abhisha3/Desktop/Projects/Excel/Gurliv.xlsx',
    'C:/Users/abhisha3/Desktop/Projects/Excel/Ishan.xlsx',
    'C:/Users/abhisha3/Desktop/Projects/Excel/Prateek.xlsx'
]

# Function to update master file with data from child files
def update_master():
    master_wb = openpyxl.load_workbook(master_file)
    master_ws = master_wb.active

    # Clear existing data in master sheet
    master_ws.delete_rows(1, master_ws.max_row)

    for child_file in child_files:
        child_wb = openpyxl.load_workbook(child_file)
        child_ws = child_wb[os.path.basename(child_file).split('.')[0]]

        # Copy data from child sheet to master sheet
        for row in child_ws.iter_rows(values_only=True):
            master_ws.append(row)

    master_wb.save(master_file)

# Monitor child files for changes and update master file
def monitor_files():
    last_modified_times = {file: os.path.getmtime(file) for file in child_files}

    while True:
        for file in child_files:
            current_modified_time = os.path.getmtime(file)
            if current_modified_time != last_modified_times[file]:
                print(f'{file} has been modified. Updating master file...')
                update_master()
                last_modified_times[file] = current_modified_time
        time.sleep(1)

if __name__ == '__main__':
    monitor_files()