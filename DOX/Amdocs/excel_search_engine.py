'''Steps to be followed:-
Open bash:-
Navigate to below path:- python script location which is having excel location embedded in code itself
Change the excel path in code if running in another machine
Abhishek Sceript psth: C/Users/abhisha3/Desktop/Projects/Amdocs
Run command to execute script: python excel_search_engine.py
Abhi Excel path:-  "C:\Users\abhisha3\Desktop\OSO_Automation\OSO_Search_Engine\Abhi_SE_V1.xlsx"
EXCEL_FILE_PATH = r"C:\Users\abhisha3\Desktop\OSO_Automation\OSO_Search_Engine\Abhi_SE_V1.xlsx"
'''



import openpyxl
import os
#Desktop/OSO_Automation/OSO_Search_Engine
#"C:\Users\abhisha3\Desktop\OSO_Automation\OSO_Search_Engine\Abhi_SE_V1.xlsx"
# --- Configuration ---
# IMPORTANT: Replace this with the actual path to your Excel file.
# Make sure the path uses double backslashes (\\) or forward slashes (/)
# For example: r"C:\Users\abhisha3\Desktop\OSO Automation\OSO Search Engine\Abhi_SE_V1.xlsx"
EXCEL_FILE_PATH = r"C:\Users\abhisha3\Desktop\OSO_Automation\OSO_Search_Engine\Abhi_SE_V1.xlsx"

# --- Data Loading ---
# This function loads data from a specified sheet in the Excel file.
def load_sheet_data(file_path, sheet_name):
    """
    Loads data from a specified Excel sheet into a list of dictionaries.
    Each dictionary represents a row, with keys as column headers.
    """
    data = []
    try:
        # Load the workbook from the given file path
        workbook = openpyxl.load_workbook(file_path)
        # Select the desired sheet
        sheet = workbook[sheet_name]

        # Get column headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Iterate through rows, starting from the second row (to skip headers)
        for row_index in range(2, sheet.max_row + 1):
            row_data = {}
            # Iterate through columns to build a dictionary for each row
            for col_index, header in enumerate(headers):
                cell_value = sheet.cell(row=row_index, column=col_index + 1).value
                row_data[header] = cell_value
            data.append(row_data)
        print(f"Successfully loaded {len(data)} rows from '{sheet_name}' sheet.")
    except FileNotFoundError:
        print(f"Error: Excel file not found at '{file_path}'. Please check the path.")
        return []
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
        return []
    except Exception as e:
        print(f"An error occurred while loading sheet '{sheet_name}': {e}")
        return []
    return data

# Global variables to store loaded data
sr_data = []
defect_data = []

# --- Search Logic ---
def search_data(sr_records, defect_records, query_params):
    """
    Performs search operations on SR and Defect records based on query parameters.
    Returns filtered lists of records.
    """
    filtered_sr = []
    filtered_defect = []

    # Helper function to check if a value contains a search term (case-insensitive)
    def contains_term(value, term):
        if value is None:
            return False
        return str(value).lower().find(str(term).lower()) != -1

    # Apply filters to SR records
    for record in sr_records:
        match_sr = True
        # CUSTOMER_ID search in SR_ID
        if query_params.get('customer_id') and not contains_term(record.get('SR_ID'), query_params['customer_id']):
            match_sr = False
        # OSite ID search in DETAILS, UPDATE_DETAILS
        if query_params.get('osite_id') and \
           not (contains_term(record.get('DETAILS'), query_params['osite_id']) or
                contains_term(record.get('UPDATE_DETAILS'), query_params['osite_id'])):
            match_sr = False
        # SR_ID search in SR_ID
        if query_params.get('sr_id') and not contains_term(record.get('SR_ID'), query_params['sr_id']):
            match_sr = False
        # ID (Defect ID) search in DETAILS, UPDATE_DETAILS
        if query_params.get('defect_id') and \
           not (contains_term(record.get('DETAILS'), query_params['defect_id']) or
                contains_term(record.get('UPDATE_DETAILS'), query_params['defect_id'])):
            match_sr = False
        # Search Anything in DETAILS, UPDATE_DETAILS
        if query_params.get('search_anything') and \
           not (contains_term(record.get('DETAILS'), query_params['search_anything']) or
                contains_term(record.get('UPDATE_DETAILS'), query_params['search_anything'])):
            match_sr = False

        if match_sr:
            filtered_sr.append(record)

    # Apply filters to Defect records
    for record in defect_records:
        match_defect = True
        # CUSTOMER_ID search in Name, Description
        if query_params.get('customer_id') and \
           not (contains_term(record.get('Name'), query_params['customer_id']) or
                contains_term(record.get('Description'), query_params['customer_id'])):
            match_defect = False
        # OSite ID search in Name, Description
        if query_params.get('osite_id') and \
           not (contains_term(record.get('Name'), query_params['osite_id']) or
                contains_term(record.get('Description'), query_params['osite_id'])):
            match_defect = False
        # SR_ID search in Name, Description
        if query_params.get('sr_id') and \
           not (contains_term(record.get('Name'), query_params['sr_id']) or
                contains_term(record.get('Description'), query_params['sr_id'])):
            match_defect = False
        # ID (Defect ID) search in ID
        if query_params.get('defect_id') and not contains_term(record.get('ID'), query_params['defect_id']):
            match_defect = False
        # Search Anything in Name, Description
        if query_params.get('search_anything') and \
           not (contains_term(record.get('Name'), query_params['search_anything']) or
                contains_term(record.get('Description'), query_params['search_anything'])):
            match_defect = False

        if match_defect:
            filtered_defect.append(record)

    return filtered_sr, filtered_defect

# --- Display Results ---
def display_results(sr_results, defect_results):
    """
    Prints the search results for SR and Defect records to the console.
    """
    print("\n" + "="*50)
    print("SR Search Results:")
    print("="*50)
    if sr_results:
        # Define columns to display for SR
        sr_display_cols = ["SR_ID", "CUSTOMER_ID", "RCA", "DETAILS", "UPDATE_DETAILS", "CUSTOMER_NAME"]
        # Print header
        print("{:<15} {:<15} {:<20} {:<30} {:<30} {:<25}".format(*sr_display_cols))
        print("-" * 135)
        # Print each record
        for record in sr_results:
            # Ensure values are truncated if too long for display
            row_values = []
            for col in sr_display_cols:
                value = str(record.get(col, 'N/A'))
                if col == "SR_ID" or col == "CUSTOMER_ID":
                    row_values.append(value[:14]) # Shorter for IDs
                elif col == "RCA":
                    row_values.append(value[:19]) # Shorter for RCA
                elif col == "CUSTOMER_NAME":
                    row_values.append(value[:24]) # Shorter for customer name
                else:
                    row_values.append(value[:29]) # Longer for details
            print("{:<15} {:<15} {:<20} {:<30} {:<30} {:<25}".format(*row_values))
    else:
        print("No SR records found matching your criteria.")

    print("\n" + "="*50)
    print("Defect Search Results:")
    print("="*50)
    if defect_results:
        # Define columns to display for Defect
        defect_display_cols = ["ID", "Name", "Description", "Phase", "Release"]
        # Print header
        print("{:<10} {:<25} {:<40} {:<15} {:<15}".format(*defect_display_cols))
        print("-" * 105)
        # Print each record
        for record in defect_results:
            row_values = []
            for col in defect_display_cols:
                value = str(record.get(col, 'N/A'))
                if col == "ID":
                    row_values.append(value[:9])
                elif col == "Name":
                    row_values.append(value[:24])
                elif col == "Description":
                    row_values.append(value[:39])
                else:
                    row_values.append(value[:14])
            print("{:<10} {:<25} {:<40} {:<15} {:<15}".format(*row_values))
    else:
        print("No Defect records found matching your criteria.")
    print("="*50 + "\n")

# --- Main Application Loop ---
def main():
    global sr_data, defect_data # Declare global to modify them

    print("Loading Excel data... This might take a moment for large files.")
    sr_data = load_sheet_data(EXCEL_FILE_PATH, "SR")
    defect_data = load_sheet_data(EXCEL_FILE_PATH, "Defect")

    if not sr_data and not defect_data:
        print("Exiting as no data could be loaded. Please fix the Excel path or file issues.")
        return

    while True:
        print("\n--- Amdocs OSO Search Engine (Console Version) ---")
        print("Enter search criteria (leave blank to ignore):")
        print("1. CUSTOMER_ID (searches SR_ID in SR; Name, Description in Defect)")
        print("2. OSite ID (e.g., 'OSite_12345_1') (searches DETAILS, UPDATE_DETAILS in SR; Name, Description in Defect)")
        print("3. SR_ID (searches SR_ID in SR; Name, Description in Defect)")
        print("4. Defect ID (searches DETAILS, UPDATE_DETAILS in SR; ID in Defect)")
        print("5. Search Anything (general keyword search in DETAILS, UPDATE_DETAILS in SR; Name, Description in Defect)")
        print("Type 'exit' to quit.")

        query_params = {}
        query_params['customer_id'] = input("Enter CUSTOMER_ID: ").strip() or None
        query_params['osite_id'] = input("Enter OSite ID: ").strip() or None
        query_params['sr_id'] = input("Enter SR_ID: ").strip() or None
        query_params['defect_id'] = input("Enter Defect ID: ").strip() or None
        query_params['search_anything'] = input("Enter anything else to search: ").strip() or None

        # Check for exit command
        if any(val and val.lower() == 'exit' for val in query_params.values() if val is not None):
            print("Exiting search engine. Goodbye!")
            break

        # Perform search
        print("\nSearching...")
        filtered_sr_results, filtered_defect_results = search_data(sr_data, defect_data, query_params)

        # Display results
        display_results(filtered_sr_results, filtered_defect_results)

        input("\nPress Enter to perform another search...") # Pause for user to read results

if __name__ == "__main__":
    main()
