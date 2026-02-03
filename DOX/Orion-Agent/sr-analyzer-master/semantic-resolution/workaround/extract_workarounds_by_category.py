import pandas as pd
from pathlib import Path
from collections import defaultdict

def read_past_data_files():
    """Read all Excel files from past_data folder"""
    past_data_dir = Path("past_data")
    all_data = []
    
    print("Reading past data files...")
    for file in past_data_dir.glob("*.xls"):
        try:
            df = pd.read_excel(file)
            all_data.append(df)
            print(f"✓ Read {file.name}: {len(df)} rows")
        except Exception as e:
            print(f"✗ Error reading {file.name}: {e}")
    
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nTotal rows from past data: {len(combined_df)}")
    
    return combined_df

def read_category_mappings():
    """Read category mapping files"""
    print("\nReading category mappings...")
    
    # Read midmarket amdocs issue mapping
    amdocs_df = pd.read_csv("category mapping/midmarket_amdocs_issue(Sheet1).csv")
    print(f"✓ Read Amdocs Issue Mapping: {len(amdocs_df)} rows")
    
    # Read incident resolution SLA category mapping (customer issues)
    customer_df = pd.read_csv("category mapping/incident_resolution_sla_category_and_status(Sheet1).csv")
    print(f"✓ Read Customer Issue Mapping: {len(customer_df)} rows")
    
    # Create lookup dictionaries
    amdocs_lookup = {}
    for _, row in amdocs_df.iterrows():
        key = f"{row['Tier1']}|{row['Tier2']}|{row['Tier3']}"
        amdocs_lookup[key] = {
            'guideline': str(row['Workaround comment']).strip() if pd.notna(row['Workaround comment']) else "",
            'status': str(row['Status']).strip() if pd.notna(row['Status']) else "",
            'pmr': str(row['PMR Masking']).strip() if pd.notna(row['PMR Masking']) else ""
        }
    
    customer_lookup = {}
    for _, row in customer_df.iterrows():
        key = f"{row['Tier1']}|{row['Tier2']}|{row['Tier3']}"
        customer_lookup[key] = {
            'guideline': str(row['Workaround comment']).strip() if pd.notna(row['Workaround comment']) else "",
            'status': str(row['Status']).strip() if pd.notna(row['Status']) else "",
            'pmr': str(row['PMR Masking']).strip() if pd.notna(row['PMR Masking']) else ""
        }
    
    return amdocs_lookup, customer_lookup

def extract_workarounds_by_category(past_df, amdocs_lookup, customer_lookup):
    """Extract workarounds grouped by category combinations"""
    print("\n" + "="*80)
    print("EXTRACTING WORKAROUNDS BY CATEGORY")
    print("="*80)
    
    # Filter rows with non-empty workarounds
    past_df_with_wa = past_df[
        past_df['Workaround'].notna() & 
        (past_df['Workaround'].str.strip() != '')
    ].copy()
    
    print(f"\nFound {len(past_df_with_wa)} records with workarounds out of {len(past_df)} total")
    
    # Group by category combinations
    workaround_data = defaultdict(list)
    
    for _, row in past_df_with_wa.iterrows():
        # Extract categorization fields
        resolution_t2 = str(row['Resolution Categorization']).strip() if pd.notna(row['Resolution Categorization']) else ""
        resolution_t3 = str(row['Resolution Categorization(Resolution Category Tier 3)']).strip() if pd.notna(row['Resolution Categorization(Resolution Category Tier 3)']) else ""
        sla_t1 = str(row['SLA Resolution Categorization T1']).strip() if pd.notna(row['SLA Resolution Categorization T1']) else ""
        sla_category = str(row['SLA Resolution Category']).strip() if pd.notna(row['SLA Resolution Category']) else ""
        workaround = str(row['Workaround']).strip() if pd.notna(row['Workaround']) else ""
        
        if resolution_t2 and resolution_t3 and sla_t1 and sla_category and workaround:
            # Create a unique key for this combination
            key = f"{resolution_t2}|{resolution_t3}|{sla_t1}|{sla_category}"
            
            # Store the workaround with all fields
            workaround_data[key].append({
                'resolution_t2': resolution_t2,
                'resolution_t3': resolution_t3,
                'sla_t1': sla_t1,
                'sla_category': sla_category,
                'workaround': workaround
            })
    
    print(f"Found {len(workaround_data)} unique category combinations")
    
    return workaround_data

def create_output_excel(workaround_data, amdocs_lookup, customer_lookup):
    """Create output Excel with all workarounds by category"""
    print("\n" + "="*80)
    print("CREATING OUTPUT EXCEL")
    print("="*80)
    
    output_rows = []
    
    # Process each category combination
    for key, workarounds in workaround_data.items():
        resolution_t2 = workarounds[0]['resolution_t2']
        resolution_t3 = workarounds[0]['resolution_t3']
        sla_t1 = workarounds[0]['sla_t1']
        sla_category = workarounds[0]['sla_category']
        
        # Get guideline from category mapping
        lookup_key = f"{sla_t1}|{sla_category}|{resolution_t3}"
        guideline_info = amdocs_lookup.get(lookup_key, customer_lookup.get(lookup_key, {}))
        
        guideline = guideline_info.get('guideline', '') if guideline_info else ''
        expected_status = guideline_info.get('status', '') if guideline_info else ''
        pmr_masking = guideline_info.get('pmr', '') if guideline_info else ''
        
        # Get all unique workaround comments for this combination
        unique_workarounds = list(set([w['workaround'] for w in workarounds]))
        
        # Add each unique workaround as a separate row
        for workaround_comment in unique_workarounds:
            output_rows.append({
                'Resolution Categorization (T2)': resolution_t2,
                'Resolution Category Tier 3 (T3)': resolution_t3,
                'SLA Resolution Categorization T1': sla_t1,
                'SLA Resolution Category': sla_category,
                'Workaround Comment': workaround_comment,
                'Workaround Guideline (from mapping)': guideline,
                'Expected Status': expected_status,
                'PMR Masking': pmr_masking,
                'Occurrences': len([w for w in workarounds if w['workaround'] == workaround_comment])
            })
    
    # Create DataFrame
    df = pd.DataFrame(output_rows)
    
    # Sort by occurrences (most common first)
    df = df.sort_values('Occurrences', ascending=False)
    
    print(f"Created {len(df)} rows")
    print(f"Unique combinations: {len(df[['Resolution Categorization (T2)', 'Resolution Category Tier 3 (T3)', 'SLA Resolution Categorization T1', 'SLA Resolution Category']].drop_duplicates())}")
    
    return df

def create_summary_sheet(main_df):
    """Create summary statistics"""
    print("\nCreating summary statistics...")
    
    summary_data = []
    
    # Group by category combination
    category_groups = main_df.groupby([
        'Resolution Categorization (T2)',
        'Resolution Category Tier 3 (T3)',
        'SLA Resolution Categorization T1',
        'SLA Resolution Category'
    ])
    
    for name, group in category_groups:
        summary_data.append({
            'Resolution Categorization (T2)': name[0],
            'Resolution Category Tier 3 (T3)': name[1],
            'SLA Resolution Categorization T1': name[2],
            'SLA Resolution Category': name[3],
            'Total Workarounds': group['Occurrences'].sum(),
            'Unique Workaround Variants': len(group),
            'Most Common Workaround': group.iloc[0]['Workaround Comment'][:100] + '...' if len(group.iloc[0]['Workaround Comment']) > 100 else group.iloc[0]['Workaround Comment'],
            'Workaround Guideline': group.iloc[0]['Workaround Guideline (from mapping)']
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df = summary_df.sort_values('Total Workarounds', ascending=False)
    
    return summary_df

def save_to_excel(main_df, summary_df):
    """Save to Excel file"""
    output_file = "Workaround_Comments_By_Category.xlsx"
    
    print(f"\nSaving to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Main sheet with all workarounds
        main_df.to_excel(writer, sheet_name='All Workarounds', index=False)
        
        # Summary sheet
        summary_df.to_excel(writer, sheet_name='Category Summary', index=False)
        
        # Split by SLA T1
        for sla_t1 in main_df['SLA Resolution Categorization T1'].unique():
            if pd.notna(sla_t1) and sla_t1:
                filtered_df = main_df[main_df['SLA Resolution Categorization T1'] == sla_t1]
                sheet_name = sla_t1[:31]  # Excel sheet name limit
                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"✓ Successfully created {output_file}")
    
    # Get workbook to count sheets
    import openpyxl
    wb = openpyxl.load_workbook(output_file)
    print(f"\nSheets created: {len(wb.sheetnames)}")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

def main():
    print("="*80)
    print("EXTRACT WORKAROUNDS BY CATEGORY COMBINATION")
    print("="*80)
    
    # Step 1: Read past data
    past_df = read_past_data_files()
    
    # Step 2: Read category mappings
    amdocs_lookup, customer_lookup = read_category_mappings()
    
    # Step 3: Extract workarounds by category
    workaround_data = extract_workarounds_by_category(past_df, amdocs_lookup, customer_lookup)
    
    # Step 4: Create output Excel
    main_df = create_output_excel(workaround_data, amdocs_lookup, customer_lookup)
    
    # Step 5: Create summary
    summary_df = create_summary_sheet(main_df)
    
    # Step 6: Save to Excel
    save_to_excel(main_df, summary_df)
    
    print("\n" + "="*80)
    print("COMPLETE!")
    print("="*80)
    print("\nOutput file: Workaround_Comments_By_Category.xlsx")
    print(f"Total workaround entries: {len(main_df)}")
    print(f"Unique category combinations: {len(summary_df)}")
    
    print("\n" + "="*80)
    print("TOP 10 MOST COMMON CATEGORY COMBINATIONS")
    print("="*80)
    for idx, row in summary_df.head(10).iterrows():
        print(f"\n{idx+1}. {row['SLA Resolution Categorization T1']} → {row['Resolution Category Tier 3 (T3)']}")
        print(f"   Resolution T2: {row['Resolution Categorization (T2)']}")
        print(f"   SLA Category: {row['SLA Resolution Category']}")
        print(f"   Total occurrences: {row['Total Workarounds']}")
        print(f"   Unique variants: {row['Unique Workaround Variants']}")
        print(f"   Guideline: {row['Workaround Guideline']}")

if __name__ == "__main__":
    main()



