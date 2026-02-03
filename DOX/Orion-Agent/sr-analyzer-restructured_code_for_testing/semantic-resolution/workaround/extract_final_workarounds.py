import pandas as pd
from pathlib import Path
from collections import defaultdict
import re

def read_past_data_files():
    """Read all Excel files from past_data folder"""
    past_data_dir = Path("past_data")
    all_data = []
    
    print("Reading past data files...")
    for file in past_data_dir.glob("*.xls"):
        try:
            df = pd.read_excel(file)
            all_data.append(df)
            print(f"✓ Read {file.name}: {len(df)} rows")
        except Exception as e:
            print(f"✗ Error reading {file.name}: {e}")
    
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nTotal rows from past data: {len(combined_df)}")
    
    return combined_df

def is_good_comment(comment):
    """Filter for meaningful workaround comments"""
    if not comment or len(comment.strip()) < 5:
        return False
    
    # Remove very generic/empty comments
    bad_patterns = [
        r'^\.+$',  # Just dots
        r'^-+$',   # Just dashes
        r'^\s*$',  # Empty/whitespace
    ]
    
    for pattern in bad_patterns:
        if re.match(pattern, comment.strip()):
            return False
    
    return True

def read_category_mappings():
    """Read category mapping files"""
    print("\nReading category mappings...")
    
    # Read midmarket amdocs issue mapping
    amdocs_df = pd.read_csv("category mapping/midmarket_amdocs_issue(Sheet1).csv")
    print(f"✓ Read Amdocs Issue Mapping: {len(amdocs_df)} rows")
    
    # Read incident resolution SLA category mapping (customer issues)
    customer_df = pd.read_csv("category mapping/incident_resolution_sla_category_and_status(Sheet1).csv")
    print(f"✓ Read Customer Issue Mapping: {len(customer_df)} rows")
    
    # Create lookup dictionaries
    category_lookup = {}
    
    for _, row in amdocs_df.iterrows():
        key = f"{row['Tier1']}|{row['Tier2']}|{row['Tier3']}"
        category_lookup[key] = {
            'guideline': str(row['Workaround comment']).strip() if pd.notna(row['Workaround comment']) else ""
        }
    
    for _, row in customer_df.iterrows():
        key = f"{row['Tier1']}|{row['Tier2']}|{row['Tier3']}"
        category_lookup[key] = {
            'guideline': str(row['Workaround comment']).strip() if pd.notna(row['Workaround comment']) else ""
        }
    
    return category_lookup

def extract_actual_workarounds(past_df, category_lookup):
    """Extract ONLY actual workaround comments from past data"""
    print("\n" + "="*80)
    print("EXTRACTING ACTUAL WORKAROUND COMMENTS FROM PAST DATA")
    print("="*80)
    
    # Filter rows with non-empty workarounds
    past_df_with_wa = past_df[
        past_df['Workaround'].notna() & 
        (past_df['Workaround'].str.strip() != '')
    ].copy()
    
    print(f"\nFound {len(past_df_with_wa)} records with workarounds out of {len(past_df)} total")
    
    # Extract and organize
    output_rows = []
    
    for _, row in past_df_with_wa.iterrows():
        # Extract categorization fields
        resolution_t2 = str(row['Resolution Categorization']).strip() if pd.notna(row['Resolution Categorization']) else ""
        resolution_t3 = str(row['Resolution Categorization(Resolution Category Tier 3)']).strip() if pd.notna(row['Resolution Categorization(Resolution Category Tier 3)']) else ""
        sla_t1 = str(row['SLA Resolution Categorization T1']).strip() if pd.notna(row['SLA Resolution Categorization T1']) else ""
        sla_category = str(row['SLA Resolution Category']).strip() if pd.notna(row['SLA Resolution Category']) else ""
        workaround = str(row['Workaround']).strip() if pd.notna(row['Workaround']) else ""
        
        # Skip if any required field is missing or workaround is not good
        if not all([resolution_t2, resolution_t3, sla_t1, sla_category, workaround]):
            continue
        
        if not is_good_comment(workaround):
            continue
        
        # Get category mapping info
        lookup_key = f"{sla_t1}|{sla_category}|{resolution_t3}"
        category_info = category_lookup.get(lookup_key, {})
        
        output_rows.append({
            'Resolution Categorization T2': resolution_t2,
            'Resolution Category Tier 3': resolution_t3,
            'SLA Resolution Categorization T1': sla_t1,
            'SLA Resolution Category': sla_category,
            'Actual Workaround Comment': workaround,
            'Workaround Guideline': category_info.get('guideline', '')
        })
    
    print(f"Extracted {len(output_rows)} actual workaround comments")
    
    return output_rows

def create_output_excel(output_rows):
    """Create output Excel with actual workarounds"""
    print("\n" + "="*80)
    print("CREATING OUTPUT EXCEL")
    print("="*80)
    
    # Create main DataFrame
    df = pd.DataFrame(output_rows)
    
    # Count occurrences of each unique workaround in same category
    df['Occurrences'] = df.groupby([
        'Resolution Categorization T2',
        'Resolution Category Tier 3',
        'SLA Resolution Categorization T1',
        'SLA Resolution Category',
        'Actual Workaround Comment'
    ])['Actual Workaround Comment'].transform('count')
    
    # Remove duplicates, keeping occurrence count
    df = df.drop_duplicates(subset=[
        'Resolution Categorization T2',
        'Resolution Category Tier 3',
        'SLA Resolution Categorization T1',
        'SLA Resolution Category',
        'Actual Workaround Comment'
    ])
    
    # Sort by occurrences (most common first)
    df = df.sort_values('Occurrences', ascending=False)
    
    print(f"Total unique workaround comments: {len(df)}")
    
    # Create summary by category
    summary_data = []
    category_groups = df.groupby([
        'Resolution Categorization T2',
        'Resolution Category Tier 3',
        'SLA Resolution Categorization T1',
        'SLA Resolution Category'
    ])
    
    for name, group in category_groups:
        summary_data.append({
            'Resolution Categorization T2': name[0],
            'Resolution Category Tier 3': name[1],
            'SLA Resolution Categorization T1': name[2],
            'SLA Resolution Category': name[3],
            'Total Actual Workarounds': len(group),
            'Total Occurrences': group['Occurrences'].sum(),
            'Most Common Workaround': group.iloc[0]['Actual Workaround Comment'],
            'Workaround Guideline': group.iloc[0]['Workaround Guideline']
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df = summary_df.sort_values('Total Occurrences', ascending=False)
    
    return df, summary_df

def save_to_excel(main_df, summary_df):
    """Save to Excel file"""
    output_file = "Workaround_Comments_Final.xlsx"
    
    print(f"\nSaving to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Main sheet - All actual workarounds
        main_df.to_excel(writer, sheet_name='All Workarounds', index=False)
        
        # Summary sheet
        summary_df.to_excel(writer, sheet_name='Category Summary', index=False)
    
    print(f"✓ Successfully created {output_file}")
    
    # Get workbook to count sheets
    import openpyxl
    wb = openpyxl.load_workbook(output_file)
    print(f"\nSheets created: {len(wb.sheetnames)}")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

def main():
    print("="*80)
    print("EXTRACT ACTUAL WORKAROUND COMMENTS")
    print("="*80)
    print("\nColumns in output:")
    print("  1. Resolution Categorization T2")
    print("  2. Resolution Category Tier 3")
    print("  3. SLA Resolution Categorization T1")
    print("  4. SLA Resolution Category")
    print("  5. Actual Workaround Comment (from past SRs)")
    print("  6. Workaround Guideline (from category mapping)")
    print("  7. Occurrences (frequency)")
    print("="*80)
    
    # Step 1: Read past data
    past_df = read_past_data_files()
    
    # Step 2: Read category mappings
    category_lookup = read_category_mappings()
    
    # Step 3: Extract actual workarounds
    output_rows = extract_actual_workarounds(past_df, category_lookup)
    
    # Step 4: Create output Excel
    main_df, summary_df = create_output_excel(output_rows)
    
    # Step 5: Save to Excel
    save_to_excel(main_df, summary_df)
    
    print("\n" + "="*80)
    print("COMPLETE!")
    print("="*80)
    print(f"\nOutput file: Workaround_Comments_Final.xlsx")
    print(f"Total actual workaround comments: {len(main_df)}")
    print(f"Unique category combinations: {len(summary_df)}")
    
    print("\n" + "="*80)
    print("TOP 10 MOST COMMON WORKAROUNDS")
    print("="*80)
    for idx, row in main_df.head(10).iterrows():
        print(f"\n{idx+1}.")
        print(f"  Resolution T2: {row['Resolution Categorization T2']}")
        print(f"  Resolution T3: {row['Resolution Category Tier 3']}")
        print(f"  SLA T1: {row['SLA Resolution Categorization T1']}")
        print(f"  SLA Category: {row['SLA Resolution Category']}")
        print(f"  Workaround: {row['Actual Workaround Comment'][:80]}")
        print(f"  Guideline: {row['Workaround Guideline']}")
        print(f"  Occurrences: {row['Occurrences']}")

if __name__ == "__main__":
    main()



